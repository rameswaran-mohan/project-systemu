#!/usr/bin/env python3
"""Read an Excel sheet and return headers and row data."""
from __future__ import annotations

from pathlib import Path

TOOL_META = {
    "name": "read_excel_sheet",
    "tool_type": "file",
    "dependencies": ["openpyxl"],
}


def run(**kwargs) -> dict:
    path: str = kwargs.get("path", "")
    sheet_name: str = kwargs.get("sheet_name", "")
    sheet_index: int = int(kwargs.get("sheet_index", 0))

    if not path:
        return {"success": False, "headers": [], "rows": [], "row_count": 0, "error": "path is required"}

    try:
        import openpyxl

        p = Path(path).expanduser()

        if not p.exists():
            return {"success": False, "headers": [], "rows": [], "row_count": 0, "error": f"File not found: {p}"}

        wb = openpyxl.load_workbook(str(p), data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "headers": [],
                    "rows": [],
                    "row_count": 0,
                    "error": f"Sheet '{sheet_name}' not found; available: {wb.sheetnames}",
                }
            ws = wb[sheet_name]
        else:
            if sheet_index >= len(wb.sheetnames):
                return {
                    "success": False,
                    "headers": [],
                    "rows": [],
                    "row_count": 0,
                    "error": f"sheet_index {sheet_index} out of range; found {len(wb.sheetnames)} sheet(s)",
                }
            ws = wb[wb.sheetnames[sheet_index]]

        all_rows = list(ws.iter_rows(values_only=True))

        if not all_rows:
            return {"success": True, "headers": [], "rows": [], "row_count": 0, "error": None}

        headers = [str(c) if c is not None else "" for c in all_rows[0]]
        rows = [[c for c in row] for row in all_rows[1:]]

        return {"success": True, "headers": headers, "rows": rows, "row_count": len(rows), "error": None}

    except Exception as exc:
        return {"success": False, "headers": [], "rows": [], "row_count": 0, "error": str(exc)}
